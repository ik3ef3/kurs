from django.contrib.auth.models import User
from django.shortcuts import render, redirect, get_object_or_404
from .models import Menu, Dish, Order, Client, OrderDish, UserRole, Ingredient, DishIngredient
from .forms import MenuForm, OrderForm, OrderDishForm
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Prefetch
from django.http import JsonResponse
from django.db.models import Sum, Count
from datetime import datetime
from django.http import HttpResponse
from openpyxl import Workbook


def menu_list(request):
    # Получаем все меню, отсортированные по дате
    menus = Menu.objects.all().order_by('-date')
    return render(request, 'menu/list.html', {'menus': menus})


def home(request):
    return render(request, 'home.html')


def create_menu(request):
    if request.method == 'POST':
        form = MenuForm(request.POST)
        if form.is_valid():
            menu = form.save()
            # Сохраняем выбранные блюда
            dishes = form.cleaned_data['dishes']
            menu.dishes.set(dishes)
            messages.success(request, 'Меню успешно создано!')
            return redirect('menu_list')
    else:
        form = MenuForm()

    return render(request, 'menu/create.html', {'form': form})


@login_required
def order_list(request):
    # Фильтрация и сортировка
    orders = Order.objects.all().select_related('client').prefetch_related(
        Prefetch('orderdish_set', queryset=OrderDish.objects.select_related('dish'))
    )

    # Сортировка
    sort_by = request.GET.get('sort', '-created_at')
    if sort_by in ['created_at', '-created_at', 'status', '-status']:
        orders = orders.order_by(sort_by)

    # Поиск
    search_query = request.GET.get('search', '')
    if search_query:
        orders = orders.filter(client__phone__icontains=search_query)

    context = {
        'orders': orders,
        'current_sort': sort_by,
        'search_query': search_query
    }
    return render(request, 'orders/list.html', context)


def create_order(request):
    if request.method == 'POST':
        form = OrderForm(request.POST)
        if form.is_valid():
            try:
                # Получаем данные клиента
                phone = form.cleaned_data['client_phone']
                client_name = request.POST.get('client_name', '')
                client_address = request.POST.get('client_address', '')

                # Создаём или получаем клиента
                client, created = Client.objects.get_or_create(
                    phone=phone,
                    defaults={
                        'name': client_name,
                        'address': client_address
                    }
                )

                # Создаём заказ с временной суммой 0
                order = Order.objects.create(
                    client=client,
                    notes=form.cleaned_data['notes'],
                    total=0
                )

                # Добавляем блюда и считаем сумму
                total = 0
                for key in request.POST:
                    if key.startswith('dish_'):
                        index = key.split('_')[1]
                        dish_id = request.POST.get(f'dish_{index}')
                        quantity = request.POST.get(f'quantity_{index}', 1)

                        if dish_id and quantity.isdigit():
                            dish = Dish.objects.get(id=dish_id)
                            quantity = int(quantity)
                            OrderDish.objects.create(
                                order=order,
                                dish=dish,
                                quantity=quantity
                            )
                            total += dish.price * quantity

                # Обновляем сумму заказа
                order.total = total
                order.save()

                return redirect('order_list')

            except Exception as e:
                # Логируем ошибку и возвращаем пользователя на форму
                print(f"Ошибка при создании заказа: {e}")
                return render(request, 'orders/create.html', {
                    'form': form,
                    'dishes': Dish.objects.all(),
                    'error': 'Произошла ошибка при сохранении заказа'
                })

    else:
        form = OrderForm()

    return render(request, 'orders/create.html', {
        'form': form,
        'dishes': Dish.objects.all()
    })


def autocomplete_client(request):
    term = request.GET.get('term', '')
    clients = Client.objects.filter(phone__icontains=term)[:5]
    results = [{
        'phone': c.phone,
        'name': c.name,
        'address': c.address  # Добавляем адрес в ответ
    } for c in clients]
    return JsonResponse(results, safe=False)


def is_dispatcher(user):
    try:
        return user.userrole.role == 'dispatcher' or user.is_superuser
    except UserRole.DoesNotExist:
        return False


def edit_order(request, pk):
    order = get_object_or_404(Order, pk=pk)

    if request.method == 'POST':
        # Обновляем примечания
        order.notes = request.POST.get('notes', '')
        order.status = request.POST.get('status', 'new')  # Добавлено
        order.save()

        # Удаляем все существующие блюда
        order.orderdish_set.all().delete()

        # Добавляем новые блюда
        total = 0
        for key in request.POST:
            if key.startswith('dish_'):
                index = key.split('_')[1]
                dish_id = request.POST.get(f'dish_{index}')
                quantity = request.POST.get(f'quantity_{index}', 1)

                if dish_id and quantity.isdigit():
                    try:
                        dish = Dish.objects.get(id=dish_id)
                        quantity = int(quantity)
                        OrderDish.objects.create(
                            order=order,
                            dish=dish,
                            quantity=quantity
                        )
                        total += dish.price * quantity
                    except (Dish.DoesNotExist, ValueError):
                        pass

        # Обновляем сумму
        order.total = total
        order.save()
        return redirect('order_list')

    return render(request, 'orders/edit.html', {
        'order': order,
        'dishes': Dish.objects.all(),
        'status_choices': Order.STATUS_CHOICES
    })


def delete_order(request, pk):
    order = get_object_or_404(Order, pk=pk)

    if request.method == 'POST':
        order.delete()
        return redirect('order_list')

    return render(request, 'orders/delete_confirm.html', {'order': order})


def reports(request):
    return render(request, 'reports/main.html')


dishes_r = None


def popularity_report(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    global dishes_r

    if start_date and end_date:
        start = datetime.strptime(start_date, '%Y-%m-%d')
        end = datetime.strptime(end_date, '%Y-%m-%d')

        dishes_r = Dish.objects.filter(
            orderdish__order__created_at__date__range=(start, end)
        ).annotate(
            total_ordered=Sum('orderdish__quantity')
        ).order_by('-total_ordered')

    # Экспорт в Excel
    if 'export' in request.GET:
        wb = Workbook()
        ws = wb.active
        ws.append(['Блюдо', 'Количество заказов', 'Цена'])
        for dish in dishes_r:
            ws.append([dish.name, dish.total_ordered, dish.price])
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="popularity.xlsx"'
        wb.save(response)
        return response

    return render(request, 'reports/popularity.html', {
        'dishes': dishes_r,
        'start_date': start_date,
        'end_date': end_date
    })


ingredients_r = []
total_cost = 0


def requirements_report(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    global ingredients_r
    global total_cost

    if start_date and end_date:
        start = datetime.strptime(start_date, '%Y-%m-%d')
        end = datetime.strptime(end_date, '%Y-%m-%d')

        # Сбор данных
        ingredients_data = DishIngredient.objects.filter(
            dish__orderdish__order__created_at__date__range=(start, end)
        ).values('ingredient').annotate(
            total_grams=Sum('quantity') * Sum('dish__orderdish__quantity')
        )
        total_cost = 0
        ingredients_r = []
        for item in ingredients_data:
            ingredient = Ingredient.objects.get(pk=item['ingredient'])
            kg = item['total_grams'] / 1000
            cost = kg * float(ingredient.price)
            ingredients_r.append({
                'name': ingredient.name,
                'quantity': round(kg, 2),
                'price': ingredient.price,
                'cost': round(cost, 2),
                'unit': 'кг'
            })
            total_cost += cost

    # Экспорт в Excel
    if 'export' in request.GET:
        wb = Workbook()
        ws = wb.active
        ws.append(['Ингредиент', 'Количество (кг)', 'Цена за кг', 'Стоимость'])
        for item in ingredients_r:
            ws.append([item['name'], item['quantity'], item['price'], item['cost']])
        ws.append(['Итого', '', '', total_cost])
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="requirements.xlsx"'
        wb.save(response)
        return response

    return render(request, 'reports/requirements.html', {
        'ingredients': ingredients_r,
        'total_cost': round(total_cost, 2),
        'start_date': start_date,
        'end_date': end_date
    })


revenue = 0
costs = 0
profit = 0


def profit_report(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    global revenue
    global costs
    global profit

    if start_date and end_date:
        start = datetime.strptime(start_date, '%Y-%m-%d')
        end = datetime.strptime(end_date, '%Y-%m-%d')

        # Выручка
        revenue = Order.objects.filter(
            created_at__date__range=(start, end)
        ).aggregate(total=Sum('total'))['total'] or 0

        # Затраты
        requirements_data = DishIngredient.objects.filter(
            dish__orderdish__order__created_at__date__range=(start, end)
        ).values('ingredient').annotate(
            total_grams=Sum('quantity') * Sum('dish__orderdish__quantity')
        )
        costs = 0
        for item in requirements_data:
            ingredient = Ingredient.objects.get(pk=item['ingredient'])
            kg = item['total_grams'] / 1000
            costs += kg * float(ingredient.price)

        profit = float(revenue) - costs

    # Экспорт в Excel
    if 'export' in request.GET:
        wb = Workbook()
        ws = wb.active
        ws.append(['Показатель', 'Сумма'])
        ws.append(['Выручка', revenue])
        ws.append(['Затраты', costs])
        ws.append(['Прибыль', profit])
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="profit.xlsx"'
        wb.save(response)
        return response

    return render(request, 'reports/profit.html', {
        'revenue': round(revenue, 2),
        'costs': round(costs, 2),
        'profit': round(profit, 2),
        'start_date': start_date,
        'end_date': end_date
    })
